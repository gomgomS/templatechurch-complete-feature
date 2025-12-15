import sys
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import calendar
from datetime import date, timedelta

sys.path.append("pytavia_core")
sys.path.append("pytavia_modules")
sys.path.append("pytavia_settings")
sys.path.append("pytavia_stdlib")
sys.path.append("pytavia_storage")
sys.path.append("pytavia_modules/view")

from flask import render_template, make_response, request
import pdfkit
from pytavia_core import database, config


class view_triwulan:

    def __init__(self):
        pass
    # end def

    def html(self):
        return render_template("admin/triwulan.html")
    # end def

    def html_triwulan(self):
        return render_template("admin/triwulan.html")
    # end def

    def html_triwulan_detail(self, year: int, quarter: int):
        q = 1 if quarter not in (1, 2, 3, 4) else quarter
        roman = {1: "I", 2: "II", 3: "III", 4: "IV"}[q]
        month_range = {1: "Jan – Mar", 2: "Apr – Jun", 3: "Jul – Sep", 4: "Okt – Des"}[q]
        quarter_label = f"Triwulan {roman}"

        # Compute first 12 Saturdays within the quarter
        first_month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        months_span = [first_month, first_month + 1, first_month + 2]
        start_date = date(year, months_span[0], 1)
        last_day = calendar.monthrange(year, months_span[-1])[1]
        end_date = date(year, months_span[-1], last_day)

        # Walk to first Saturday
        cur = start_date
        while cur.weekday() != 5 and cur <= end_date:
            cur += timedelta(days=1)
        saturdays = []
        while cur <= end_date and len(saturdays) < 13:
            saturdays.append(cur)
            cur += timedelta(days=7)
        # Ensure 13 entries (pad if ever needed)
        indo_months = ["jan","feb","mar","apr","mei","jun","jul","agu","sep","okt","nov","des"]
        saturday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in saturdays]
        if len(saturday_dates) < 13:
            saturday_dates += ["-"] * (13 - len(saturday_dates))
        elif len(saturday_dates) > 13:
            saturday_dates = saturday_dates[:13]

        # Compute first 12 Wednesdays within the quarter (Rabu Malam)
        cur_wed = start_date
        while cur_wed.weekday() != 2 and cur_wed <= end_date:  # 2 = Wednesday
            cur_wed += timedelta(days=1)
        wednesdays = []
        while cur_wed <= end_date and len(wednesdays) < 13:
            wednesdays.append(cur_wed)
            cur_wed += timedelta(days=7)
        wednesday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in wednesdays]
        if len(wednesday_dates) < 13:
            wednesday_dates += ["-"] * (13 - len(wednesday_dates))
        elif len(wednesday_dates) > 13:
            wednesday_dates = wednesday_dates[:13]

        # Compute first 13 Fridays within the quarter (Vesper)
        cur_fri = start_date
        while cur_fri.weekday() != 4 and cur_fri <= end_date:  # 4 = Friday
            cur_fri += timedelta(days=1)
        fridays = []
        while cur_fri <= end_date and len(fridays) < 13:
            fridays.append(cur_fri)
            cur_fri += timedelta(days=7)
        friday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in fridays]
        if len(friday_dates) < 13:
            friday_dates += ["-"] * (13 - len(friday_dates))
        elif len(friday_dates) > 13:
            friday_dates = friday_dates[:13]

        def cycle(names, n=13):
            if not names:
                return [""] * n
            return [names[i % len(names)] for i in range(n)]
        # Build data from DB participant schedules
        try:
            mgd = database.get_db_conn(config.mainDB)
            cursor = mgd.db_participant_schedule.find(
                {"year": int(year), "quarter": int(q)},
                {"_id": 0, "week_index": 1, "participant": 1},
            )
            week_to_part = {}
            for doc in cursor:
                try:
                    wi = int(doc.get("week_index", 0))
                    if 1 <= wi <= 13:
                        week_to_part[wi] = doc.get("participant", {}) or {}
                except Exception:
                    continue
        except Exception as e:
            week_to_part = {}
            try:
                print("[view_triwulan.html_detail] fetch schedules error:", str(e))
            except Exception:
                pass

        def build_series(tab_key: str, field_key: str) -> list:
            values = []
            for idx in range(1, 14):
                part = week_to_part.get(idx, {})
                tab = part.get(tab_key, {}) if isinstance(part, dict) else {}
                val = ""
                if isinstance(tab, dict):
                    val = tab.get(field_key, "") or ""
                values.append(val)
            return values

        data_khotbah = {
            "Pelayanan": build_series("khotbah", "pelayanan"),
            "Protokol": build_series("khotbah", "protokol"),
            "Pendamping": build_series("khotbah", "pendamping"),
            "Cerita Anak-anak": build_series("khotbah", "cerita_anak_anak"),
            "Pemimpin Lagu": build_series("khotbah", "pemimpin_lagu"),
            "Pianist": build_series("khotbah", "pianist"),
            "Backing Vocal": build_series("khotbah", "backing_vocal"),
            "Khotbah & SS": build_series("khotbah", "khotbah_dan_ss"),
            "Lagu Pujian": build_series("khotbah", "lagu_pujian"),
            "Diakon/Diakones": build_series("khotbah", "diakon_diakones"),
            "Penerima Tamu": build_series("khotbah", "penerima_tamu"),
        }
        data_ss = {
            "Pemimpin Acara": build_series("sekolah_sabat", "pemimpin_acara"),
            "Ayat Hafalan & Doa": build_series("sekolah_sabat", "ayat_hafalan_dan_doa"),
            "Berita Mission": build_series("sekolah_sabat", "berita_mission"),
            "Pemimpin Lagu": build_series("sekolah_sabat", "pemimpin_lagu"),
            "Pianis": build_series("sekolah_sabat", "pianis"),
            "Lagu Pujian": build_series("sekolah_sabat", "lagu_pujian"),
            "Pelayanan Perorangan": build_series("sekolah_sabat", "pelayanan_perorangan"),
            "Rumah Tangga": build_series("sekolah_sabat", "rumah_tangga"),
        }
        data_rabu = {
            "Renungan": build_series("rabu_malam", "renungan"),
            "Protokol": build_series("rabu_malam", "protokol"),
            "Pianis": build_series("rabu_malam", "pianis"),
            "Pemimpin Lagu": build_series("rabu_malam", "pemimpin_lagu"),
            "Lagu Pujian": build_series("rabu_malam", "lagu_pujian"),
        }
        data_vesper = {
            "Renungan": build_series("vesper", "renungan"),
            "Protokol": build_series("vesper", "protokol"),
            "Pianis": build_series("vesper", "pianis"),
            "Pemimpin Lagu": build_series("vesper", "pemimpin_lagu"),
            "Lagu Pujian": build_series("vesper", "lagu_pujian"),
        }

        return render_template(
            "admin/triwulan_detail.html",
            year=year,
            quarter=q,
            quarter_label=quarter_label,
            month_range=month_range,
            saturday_dates=saturday_dates,
            wednesday_dates=wednesday_dates,
            friday_dates=friday_dates,
            data_khotbah=data_khotbah,
            data_ss=data_ss,
            data_rabu=data_rabu,
            data_vesper=data_vesper,
        )
    # end def

    def export_triwulan_csv(self, year: int, quarter: int, kind: str):
        # Reuse computation and "DB" dump
        self.html_triwulan_detail(year, quarter)
        q = 1 if quarter not in (1, 2, 3, 4) else quarter
        first_month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        months_span = [first_month, first_month + 1, first_month + 2]
        start_date = date(year, months_span[0], 1)
        last_day = calendar.monthrange(year, months_span[-1])[1]
        end_date = date(year, months_span[-1], last_day)
        indo_months = ["jan","feb","mar","apr","mei","jun","jul","agu","sep","okt","nov","des"]
        cur = start_date
        while cur.weekday() != 5 and cur <= end_date:
            cur += timedelta(days=1)
        saturdays = []
        while cur <= end_date and len(saturdays) < 13:
            saturdays.append(cur)
            cur += timedelta(days=7)
        saturday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in saturdays]
        if len(saturday_dates) < 13: saturday_dates += ["-"] * (13 - len(saturday_dates))
        if len(saturday_dates) > 13: saturday_dates = saturday_dates[:13]
        wednesdays = [sat - timedelta(days=3) for sat in saturdays]
        wednesday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in wednesdays]
        if len(wednesday_dates) < 13: wednesday_dates += ["-"] * (13 - len(wednesday_dates))
        if len(wednesday_dates) > 13: wednesday_dates = wednesday_dates[:13]
        cur_fri = start_date
        while cur_fri.weekday() != 4 and cur_fri <= end_date:
            cur_fri += timedelta(days=1)
        fridays = []
        while cur_fri <= end_date and len(fridays) < 13:
            fridays.append(cur_fri)
            cur_fri += timedelta(days=7)
        friday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in fridays]
        if len(friday_dates) < 13: friday_dates += ["-"] * (13 - len(friday_dates))
        if len(friday_dates) > 13: friday_dates = friday_dates[:13]
        try:
            mgd = database.get_db_conn(config.mainDB)
            cursor = mgd.db_participant_schedule.find(
                {"year": int(year), "quarter": int(q)},
                {"_id": 0, "week_index": 1, "participant": 1},
            )
            week_to_part = {}
            for doc in cursor:
                wi = int(doc.get("week_index", 0))
                if 1 <= wi <= 13:
                    week_to_part[wi] = doc.get("participant", {}) or {}
        except Exception:
            week_to_part = {}

        def _series(tab_key: str, field_key: str) -> list:
            out = []
            for idx in range(1, 14):
                part = week_to_part.get(idx, {})
                tab = part.get(tab_key, {}) if isinstance(part, dict) else {}
                out.append((tab.get(field_key, "") if isinstance(tab, dict) else "") or "")
            return out

        data_khotbah = {
            "Pelayanan": _series("khotbah", "pelayanan"),
            "Protokol": _series("khotbah", "protokol"),
            "Pendamping": _series("khotbah", "pendamping"),
            "Cerita Anak-anak": _series("khotbah", "cerita_anak_anak"),
            "Pemimpin Lagu": _series("khotbah", "pemimpin_lagu"),
            "Pianist": _series("khotbah", "pianist"),
            "Backing Vocal": _series("khotbah", "backing_vocal"),
            "Khotbah & SS": _series("khotbah", "khotbah_dan_ss"),
            "Lagu Pujian": _series("khotbah", "lagu_pujian"),
            "Diakon/Diakones": _series("khotbah", "diakon_diakones"),
            "Penerima Tamu": _series("khotbah", "penerima_tamu"),
        }
        data_ss = {
            "Pemimpin Acara": _series("sekolah_sabat", "pemimpin_acara"),
            "Ayat Hafalan & Doa": _series("sekolah_sabat", "ayat_hafalan_dan_doa"),
            "Berita Mission": _series("sekolah_sabat", "berita_mission"),
            "Pemimpin Lagu": _series("sekolah_sabat", "pemimpin_lagu"),
            "Pianis": _series("sekolah_sabat", "pianis"),
            "Lagu Pujian": _series("sekolah_sabat", "lagu_pujian"),
            "Pelayanan Perorangan": _series("sekolah_sabat", "pelayanan_perorangan"),
            "Rumah Tangga": _series("sekolah_sabat", "rumah_tangga"),
        }
        data_rabu = {
            "Renungan": _series("rabu_malam", "renungan"),
            "Protokol": _series("rabu_malam", "protokol"),
            "Pianis": _series("rabu_malam", "pianis"),
            "Pemimpin Lagu": _series("rabu_malam", "pemimpin_lagu"),
            "Lagu Pujian": _series("rabu_malam", "lagu_pujian"),
        }
        data_vesper = {
            "Renungan": _series("vesper", "renungan"),
            "Protokol": _series("vesper", "protokol"),
            "Pianis": _series("vesper", "pianis"),
            "Pemimpin Lagu": _series("vesper", "pemimpin_lagu"),
            "Lagu Pujian": _series("vesper", "lagu_pujian"),
        }

        kind = (kind or "").lower()
        if kind == "khotbah":
            title = "Khotbah"
            rows = data_khotbah
            dates = saturday_dates
            first_col = "Sabat ke"
        elif kind in ("ss", "sekolah-sabat", "sekolah_sabat"):
            title = "Sekolah Sabat"
            rows = data_ss
            dates = saturday_dates
            first_col = "Sabat ke"
        elif kind == "vesper":
            title = "Vesper"
            rows = data_vesper
            dates = friday_dates
            first_col = "Jumat ke"
        else:
            title = "Rabu Malam"
            rows = data_rabu
            dates = wednesday_dates
            first_col = "Rabu ke"

        output = io.StringIO()
        writer = csv.writer(output)
        header = [first_col] + [str(i) for i in range(1, 14)]
        writer.writerow([f"{title} Triwulan {q} {year}"])
        writer.writerow([])
        writer.writerow(header)
        writer.writerow(["Tanggal"] + dates)
        for role, values in rows.items():
            writer.writerow([role] + values)
        csv_data = output.getvalue()
        output.close()

        resp = make_response(csv_data)
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        resp.headers["Content-Disposition"] = f"attachment; filename={title.replace(' ','_').lower()}_triwulan_{q}_{year}.csv"
        return resp
    # end def

    def export_triwulan_xlsx(self, year: int, quarter: int, kind: str):
        # Prepare datasets
        q = 1 if quarter not in (1, 2, 3, 4) else quarter
        first_month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        months_span = [first_month, first_month + 1, first_month + 2]
        start_date = date(year, months_span[0], 1)
        last_day = calendar.monthrange(year, months_span[-1])[1]
        end_date = date(year, months_span[-1], last_day)
        indo_months = ["jan","feb","mar","apr","mei","jun","jul","agu","sep","okt","nov","des"]
        cur = start_date
        while cur.weekday() != 5 and cur <= end_date:
            cur += timedelta(days=1)
        saturdays = []
        while cur <= end_date and len(saturdays) < 13:
            saturdays.append(cur)
            cur += timedelta(days=7)
        saturday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in saturdays]
        if len(saturday_dates) < 13: saturday_dates += ["-"] * (13 - len(saturday_dates))
        if len(saturday_dates) > 13: saturday_dates = saturday_dates[:13]
        wednesdays = [sat - timedelta(days=3) for sat in saturdays]
        wednesday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in wednesdays]
        if len(wednesday_dates) < 13: wednesday_dates += ["-"] * (13 - len(wednesday_dates))
        if len(wednesday_dates) > 13: wednesday_dates = wednesday_dates[:13]
        cur_fri = start_date
        while cur_fri.weekday() != 4 and cur_fri <= end_date:
            cur_fri += timedelta(days=1)
        fridays = []
        while cur_fri <= end_date and len(fridays) < 13:
            fridays.append(cur_fri)
            cur_fri += timedelta(days=7)
        friday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in fridays]
        if len(friday_dates) < 13: friday_dates += ["-"] * (13 - len(friday_dates))
        if len(friday_dates) > 13: friday_dates = friday_dates[:13]

        try:
            mgd = database.get_db_conn(config.mainDB)
            cursor = mgd.db_participant_schedule.find(
                {"year": int(year), "quarter": int(q)},
                {"_id": 0, "week_index": 1, "participant": 1},
            )
            week_to_part = {}
            for doc in cursor:
                wi = int(doc.get("week_index", 0))
                if 1 <= wi <= 13:
                    week_to_part[wi] = doc.get("participant", {}) or {}
        except Exception:
            week_to_part = {}

        def _series(tab_key: str, field_key: str) -> list:
            out = []
            for idx in range(1, 14):
                part = week_to_part.get(idx, {})
                tab = part.get(tab_key, {}) if isinstance(part, dict) else {}
                out.append((tab.get(field_key, "") if isinstance(tab, dict) else "") or "")
            return out

        data_khotbah = {
            "Pelayanan": _series("khotbah", "pelayanan"),
            "Protokol": _series("khotbah", "protokol"),
            "Pendamping": _series("khotbah", "pendamping"),
            "Cerita Anak-anak": _series("khotbah", "cerita_anak_anak"),
            "Pemimpin Lagu": _series("khotbah", "pemimpin_lagu"),
            "Pianist": _series("khotbah", "pianist"),
            "Backing Vocal": _series("khotbah", "backing_vocal"),
            "Khotbah & SS": _series("khotbah", "khotbah_dan_ss"),
            "Lagu Pujian": _series("khotbah", "lagu_pujian"),
            "Diakon/Diakones": _series("khotbah", "diakon_diakones"),
            "Penerima Tamu": _series("khotbah", "penerima_tamu"),
        }
        data_ss = {
            "Pemimpin Acara": _series("sekolah_sabat", "pemimpin_acara"),
            "Ayat Hafalan & Doa": _series("sekolah_sabat", "ayat_hafalan_dan_doa"),
            "Berita Mission": _series("sekolah_sabat", "berita_mission"),
            "Pemimpin Lagu": _series("sekolah_sabat", "pemimpin_lagu"),
            "Pianis": _series("sekolah_sabat", "pianis"),
            "Lagu Pujian": _series("sekolah_sabat", "lagu_pujian"),
            "Pelayanan Perorangan": _series("sekolah_sabat", "pelayanan_perorangan"),
            "Rumah Tangga": _series("sekolah_sabat", "rumah_tangga"),
        }
        data_rabu = {
            "Renungan": _series("rabu_malam", "renungan"),
            "Protokol": _series("rabu_malam", "protokol"),
            "Pianis": _series("rabu_malam", "pianis"),
            "Pemimpin Lagu": _series("rabu_malam", "pemimpin_lagu"),
            "Lagu Pujian": _series("rabu_malam", "lagu_pujian"),
        }
        data_vesper = {
            "Renungan": _series("vesper", "renungan"),
            "Protokol": _series("vesper", "protokol"),
            "Pianis": _series("vesper", "pianis"),
            "Pemimpin Lagu": _series("vesper", "pemimpin_lagu"),
            "Lagu Pujian": _series("vesper", "lagu_pujian"),
        }

        kind = (kind or "").lower()
        if kind == "khotbah":
            title = "Khotbah"
            rows = data_khotbah
            dates = saturday_dates
            first_col = "Sabat ke"
        elif kind in ("ss", "sekolah-sabat", "sekolah_sabat"):
            title = "Sekolah Sabat"
            rows = data_ss
            dates = saturday_dates
            first_col = "Sabat ke"
        elif kind == "vesper":
            title = "Vesper"
            rows = data_vesper
            dates = friday_dates
            first_col = "Jumat ke"
        else:
            title = "Rabu Malam"
            rows = data_rabu
            dates = wednesday_dates
            first_col = "Rabu ke"

        wb = Workbook()
        ws = wb.active
        ws.title = title

        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left = Alignment(horizontal="left", vertical="center", wrap_text=False)
        title_font = Font(size=16, bold=True)
        subtitle_font = Font(size=12, bold=True)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
        date_fill = PatternFill("solid", fgColor="FFE699")    # darker yellow

        # Header lines (merged across A..N) for 13 weeks (+1 first col) => 14 columns
        ws.merge_cells("A1:N1")
        ws.merge_cells("A2:N2")
        ws.merge_cells("A3:N3")
        ws["A1"] = "JADWAL KEBAKTIAN"
        ws["A2"] = "GMAHK JISDAC, JAKARTA"
        roman = {1: "I", 2: "II", 3: "III", 4: "IV"}[q]
        month_range = {1: "Januari - Maret", 2: "April - Juni", 3: "Juli - September", 4: "Oktober - Desember"}[q]
        ws["A3"] = f"Triwulan ke {roman} ({month_range} {year})"
        for r in (1, 2, 3):
            c = ws[f"A{r}"]
            c.font = title_font if r == 1 else subtitle_font
            c.alignment = center

        start_row = 5
        headers = [first_col] + [str(i) for i in range(1, 14)]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=text)
            cell.font = subtitle_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border
        ws.cell(row=start_row + 1, column=1, value="Tanggal").alignment = left
        ws.cell(row=start_row + 1, column=1).fill = date_fill
        ws.cell(row=start_row + 1, column=1).font = Font(bold=True)
        ws.cell(row=start_row + 1, column=1).border = border
        for i, d in enumerate(dates, start=2):
            cell = ws.cell(row=start_row + 1, column=i, value=d)
            cell.alignment = center
            cell.fill = date_fill
            cell.border = border

        current_row = start_row + 2
        for role, values in rows.items():
            cell_role = ws.cell(row=current_row, column=1, value=role)
            cell_role.alignment = left
            cell_role.border = border
            for i, v in enumerate(values, start=2):
                cell = ws.cell(row=current_row, column=i, value=v)
                cell.alignment = center
                cell.border = border
            current_row += 1

        ws.column_dimensions["A"].width = 26
        for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N"]:
            ws.column_dimensions[col_letter].width = 18

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        resp = make_response(stream.read())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{title.replace(' ','_').lower()}_triwulan_{q}_{year}.xlsx"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return resp
    # end def

    def export_triwulan_pdf(self, year: int, quarter: int, kind: str):
        q = 1 if quarter not in (1, 2, 3, 4) else quarter
        first_month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        months_span = [first_month, first_month + 1, first_month + 2]
        start_date = date(year, months_span[0], 1)
        last_day = calendar.monthrange(year, months_span[-1])[1]
        end_date = date(year, months_span[-1], last_day)
        indo_months_short = ["jan","feb","mar","apr","mei","jun","jul","agu","sep","okt","nov","des"]
        cur = start_date
        while cur.weekday() != 5 and cur <= end_date:
            cur += timedelta(days=1)
        saturdays = []
        while len(saturdays) < 13:
            saturdays.append(cur)
            cur += timedelta(days=7)
        saturday_dates = [f"{d.day} {indo_months_short[d.month-1]} {d.year}" for d in saturdays]
        if len(saturday_dates) < 13: saturday_dates += ["-"] * (13 - len(saturday_dates))
        if len(saturday_dates) > 13: saturday_dates = saturday_dates[:13]
        wednesdays = [sat - timedelta(days=3) for sat in saturdays]
        wednesday_dates = [f"{d.day} {indo_months_short[d.month-1]} {d.year}" for d in wednesdays]
        if len(wednesday_dates) < 13: wednesday_dates += ["-"] * (13 - len(wednesday_dates))
        if len(wednesday_dates) > 13: wednesday_dates = wednesday_dates[:13]
        cur_fri = start_date
        while cur_fri.weekday() != 4 and cur_fri <= end_date:
            cur_fri += timedelta(days=1)
        fridays = []
        while cur_fri <= end_date and len(fridays) < 13:
            fridays.append(cur_fri)
            cur_fri += timedelta(days=7)
        friday_dates = [f"{d.day} {indo_months_short[d.month-1]} {d.year}" for d in fridays]
        if len(friday_dates) < 13: friday_dates += ["-"] * (13 - len(friday_dates))
        if len(friday_dates) > 13: friday_dates = friday_dates[:13]
        try:
            mgd = database.get_db_conn(config.mainDB)
            cursor = mgd.db_participant_schedule.find(
                {"year": int(year), "quarter": int(q)},
                {"_id": 0, "week_index": 1, "participant": 1},
            )
            week_to_part = {}
            for doc in cursor:
                wi = int(doc.get("week_index", 0))
                if 1 <= wi <= 13:
                    week_to_part[wi] = doc.get("participant", {}) or {}
        except Exception:
            week_to_part = {}

        def _series(tab_key: str, field_key: str) -> list:
            out = []
            for idx in range(1, 14):
                part = week_to_part.get(idx, {})
                tab = part.get(tab_key, {}) if isinstance(part, dict) else {}
                out.append((tab.get(field_key, "") if isinstance(tab, dict) else "") or "")
            return out

        data_khotbah = {
            "Pelayanan": _series("khotbah", "pelayanan"),
            "Protokol": _series("khotbah", "protokol"),
            "Pendamping": _series("khotbah", "pendamping"),
            "Cerita Anak-anak": _series("khotbah", "cerita_anak_anak"),
            "Pemimpin Lagu": _series("khotbah", "pemimpin_lagu"),
            "Pianist": _series("khotbah", "pianist"),
            "Backing Vocal": _series("khotbah", "backing_vocal"),
            "Khotbah & SS": _series("khotbah", "khotbah_dan_ss"),
            "Lagu Pujian": _series("khotbah", "lagu_pujian"),
            "Diakon/Diakones": _series("khotbah", "diakon_diakones"),
            "Penerima Tamu": _series("khotbah", "penerima_tamu"),
        }
        data_ss = {
            "Pemimpin Acara": _series("sekolah_sabat", "pemimpin_acara"),
            "Ayat Hafalan & Doa": _series("sekolah_sabat", "ayat_hafalan_dan_doa"),
            "Berita Mission": _series("sekolah_sabat", "berita_mission"),
            "Pemimpin Lagu": _series("sekolah_sabat", "pemimpin_lagu"),
            "Pianis": _series("sekolah_sabat", "pianis"),
            "Lagu Pujian": _series("sekolah_sabat", "lagu_pujian"),
            "Pelayanan Perorangan": _series("sekolah_sabat", "pelayanan_perorangan"),
            "Rumah Tangga": _series("sekolah_sabat", "rumah_tangga"),
        }
        data_rabu = {
            "Renungan": _series("rabu_malam", "renungan"),
            "Protokol": _series("rabu_malam", "protokol"),
            "Pianis": _series("rabu_malam", "pianis"),
            "Pemimpin Lagu": _series("rabu_malam", "pemimpin_lagu"),
            "Lagu Pujian": _series("rabu_malam", "lagu_pujian"),
        }
        data_vesper = {
            "Renungan": _series("vesper", "renungan"),
            "Protokol": _series("vesper", "protokol"),
            "Pianis": _series("vesper", "pianis"),
            "Pemimpin Lagu": _series("vesper", "pemimpin_lagu"),
            "Lagu Pujian": _series("vesper", "lagu_pujian"),
        }

        kind_l = (kind or "").lower()
        roman = {1: "I", 2: "II", 3: "III", 4: "IV"}[q]
        month_full = {1: "Januari - Maret", 2: "April - Juni", 3: "Juli - September", 4: "Oktober - Desember"}[q]

        if kind_l == "all":
            html = render_template(
                "admin/export_triwulan_all.html",
                header_1="JADWAL KEBAKTIAN",
                header_2="GMAHK JISDAC, JAKARTA",
                header_3=f"Triwulan ke {roman} ({month_full} {year})",
                kh_first_col="Sabat ke",
                kh_dates=saturday_dates,
                kh_rows=data_khotbah,
                ss_dates=saturday_dates,
                ss_rows=data_ss,
                rabu_dates=wednesday_dates,
                rabu_rows=data_rabu,
                vesper_dates=friday_dates,
                vesper_rows=data_vesper,
            )
            filename = f"jadwal_kebaktian_triwulan_{q}_{year}_all.pdf"
        else:
            if kind_l == "khotbah":
                title = "Khotbah"
                rows = data_khotbah
                dates = saturday_dates
                first_col = "Sabat ke"
            elif kind_l in ("ss", "sekolah-sabat", "sekolah_sabat"):
                title = "Sekolah Sabat"
                rows = data_ss
                dates = saturday_dates
                first_col = "Sabat ke"
            elif kind_l == "vesper":
                title = "Vesper"
                rows = data_vesper
                dates = friday_dates
                first_col = "Jumat ke"
            else:
                title = "Rabu Malam"
                rows = data_rabu
                dates = wednesday_dates
                first_col = "Rabu ke"

            html = render_template(
                "admin/export_triwulan_table.html",
                header_1="JADWAL KEBAKTIAN",
                header_2="GMAHK JISDAC, JAKARTA",
                header_3=f"Triwulan ke {roman} ({month_full} {year})",
                title=title,
                first_col=first_col,
                dates=dates,
                rows=rows,
            )
            filename = f"{title.replace(' ','_').lower()}_triwulan_{q}_{year}.pdf"

        page_size = (request.args.get("page_size") or "A4").upper()
        orientation = request.args.get("orientation") or "Landscape"
        page_width = request.args.get("page_width")
        page_height = request.args.get("page_height")
        margin_top = request.args.get("margin_top") or "10mm"
        margin_right = request.args.get("margin_right") or "10mm"
        margin_bottom = request.args.get("margin_bottom") or "10mm"
        margin_left = request.args.get("margin_left") or "10mm"
        options = {
            "orientation": orientation,
            "encoding": "UTF-8",
            "margin-top": margin_top,
            "margin-right": margin_right,
            "margin-bottom": margin_bottom,
            "margin-left": margin_left,
        }
        if page_width and page_height:
            options["page-width"] = page_width
            options["page-height"] = page_height
        else:
            options["page-size"] = page_size
        pdf_bytes = pdfkit.from_string(html, False, options=options)
        resp = make_response(pdf_bytes)
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return resp
    # end def

    def export_triwulan_xlsx_all(self, year: int, quarter: int):
        q = 1 if quarter not in (1, 2, 3, 4) else quarter
        first_month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        months_span = [first_month, first_month + 1, first_month + 2]
        start_date = date(year, months_span[0], 1)
        last_day = calendar.monthrange(year, months_span[-1])[1]
        end_date = date(year, months_span[-1], last_day)
        indo_months = ["jan","feb","mar","apr","mei","jun","jul","agu","sep","okt","nov","des"]
        cur = start_date
        while cur.weekday() != 5 and cur <= end_date:
            cur += timedelta(days=1)
        saturdays = []
        while cur <= end_date and len(saturdays) < 13:
            saturdays.append(cur)
            cur += timedelta(days=7)
        saturday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in saturdays]
        if len(saturday_dates) < 13: saturday_dates += ["-"] * (13 - len(saturday_dates))
        if len(saturday_dates) > 13: saturday_dates = saturday_dates[:13]
        cur_wed = start_date
        while cur_wed.weekday() != 2 and cur_wed <= end_date:
            cur_wed += timedelta(days=1)
        wednesdays = []
        while cur_wed <= end_date and len(wednesdays) < 13:
            wednesdays.append(cur_wed)
            cur_wed += timedelta(days=7)
        wednesday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in wednesdays]
        if len(wednesday_dates) < 13: wednesday_dates += ["-"] * (13 - len(wednesday_dates))
        if len(wednesday_dates) > 13: wednesday_dates = wednesday_dates[:13]
        cur_fri = start_date
        while cur_fri.weekday() != 4 and cur_fri <= end_date:
            cur_fri += timedelta(days=1)
        fridays = []
        while cur_fri <= end_date and len(fridays) < 13:
            fridays.append(cur_fri)
            cur_fri += timedelta(days=7)
        friday_dates = [f"{d.day} {indo_months[d.month-1]} {d.year}" for d in fridays]
        if len(friday_dates) < 13: friday_dates += ["-"] * (13 - len(friday_dates))
        if len(friday_dates) > 13: friday_dates = friday_dates[:13]

        try:
            mgd = database.get_db_conn(config.mainDB)
            cursor = mgd.db_participant_schedule.find(
                {"year": int(year), "quarter": int(q)},
                {"_id": 0, "week_index": 1, "participant": 1},
            )
            week_to_part = {}
            for doc in cursor:
                wi = int(doc.get("week_index", 0))
                if 1 <= wi <= 13:
                    week_to_part[wi] = doc.get("participant", {}) or {}
        except Exception:
            week_to_part = {}

        def _series(tab_key: str, field_key: str) -> list:
            out = []
            for idx in range(1, 14):
                part = week_to_part.get(idx, {})
                tab = part.get(tab_key, {}) if isinstance(part, dict) else {}
                out.append((tab.get(field_key, "") if isinstance(tab, dict) else "") or "")
            return out

        data_khotbah = {
            "Pelayanan": _series("khotbah", "pelayanan"),
            "Protokol": _series("khotbah", "protokol"),
            "Pendamping": _series("khotbah", "pendamping"),
            "Cerita Anak-anak": _series("khotbah", "cerita_anak_anak"),
            "Pemimpin Lagu": _series("khotbah", "pemimpin_lagu"),
            "Pianist": _series("khotbah", "pianist"),
            "Backing Vocal": _series("khotbah", "backing_vocal"),
            "Khotbah & SS": _series("khotbah", "khotbah_dan_ss"),
            "Lagu Pujian": _series("khotbah", "lagu_pujian"),
            "Diakon/Diakones": _series("khotbah", "diakon_diakones"),
            "Penerima Tamu": _series("khotbah", "penerima_tamu"),
        }
        data_ss = {
            "Pemimpin Acara": _series("sekolah_sabat", "pemimpin_acara"),
            "Ayat Hafalan & Doa": _series("sekolah_sabat", "ayat_hafalan_dan_doa"),
            "Berita Mission": _series("sekolah_sabat", "berita_mission"),
            "Pemimpin Lagu": _series("sekolah_sabat", "pemimpin_lagu"),
            "Pianis": _series("sekolah_sabat", "pianis"),
            "Lagu Pujian": _series("sekolah_sabat", "lagu_pujian"),
            "Pelayanan Perorangan": _series("sekolah_sabat", "pelayanan_perorangan"),
            "Rumah Tangga": _series("sekolah_sabat", "rumah_tangga"),
        }
        data_rabu = {
            "Renungan": _series("rabu_malam", "renungan"),
            "Protokol": _series("rabu_malam", "protokol"),
            "Pianis": _series("rabu_malam", "pianis"),
            "Pemimpin Lagu": _series("rabu_malam", "pemimpin_lagu"),
            "Lagu Pujian": _series("rabu_malam", "lagu_pujian"),
        }
        data_vesper = {
            "Renungan": _series("vesper", "renungan"),
            "Protokol": _series("vesper", "protokol"),
            "Pianis": _series("vesper", "pianis"),
            "Pemimpin Lagu": _series("vesper", "pemimpin_lagu"),
            "Lagu Pujian": _series("vesper", "lagu_pujian"),
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "Jadwal Kebaktian"

        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left = Alignment(horizontal="left", vertical="center", wrap_text=False)
        title_font = Font(size=16, bold=True)
        subtitle_font = Font(size=12, bold=True)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="FFF2CC")
        date_fill = PatternFill("solid", fgColor="FFE699")

        ws.merge_cells("A1:N1")
        ws.merge_cells("A2:N2")
        ws.merge_cells("A3:N3")
        ws["A1"] = "JADWAL KEBAKTIAN"
        ws["A2"] = "GMAHK JISDAC, JAKARTA"
        roman = {1: "I", 2: "II", 3: "III", 4: "IV"}[q]
        month_range_full = {1: "Januari - Maret", 2: "April - Juni", 3: "Juli - September", 4: "Oktober - Desember"}[q]
        ws["A3"] = f"Triwulan ke {roman} ({month_range_full} {year})"
        for r in (1, 2, 3):
            c = ws[f"A{r}"]
            c.font = title_font if r == 1 else subtitle_font
            c.alignment = center

        def write_table(start_row: int, first_col_name: str, dates: list, rows_map: dict, include_header: bool = True, date_label: str = "Tanggal") -> int:
            row_idx = start_row
            if include_header:
                headers = [first_col_name] + [str(i) for i in range(1, 14)]
                for col_idx, text in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=text)
                    cell.font = subtitle_font
                    cell.alignment = center
                    cell.fill = header_fill
                    cell.border = border
                row_idx += 1
            ws.cell(row=row_idx, column=1, value=date_label).alignment = left
            ws.cell(row=row_idx, column=1).fill = date_fill
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).border = border
            for i, d in enumerate(dates, start=2):
                cell = ws.cell(row=row_idx, column=i, value=d)
                cell.alignment = center
                cell.fill = date_fill
                cell.border = border
            row_idx += 1
            for role, values in rows_map.items():
                cell_role = ws.cell(row=row_idx, column=1, value=role)
                cell_role.alignment = left
                cell_role.border = border
                for i, v in enumerate(values, start=2):
                    cell = ws.cell(row=row_idx, column=i, value=v)
                    cell.alignment = center
                    cell.border = border
                row_idx += 1
            return row_idx

        current = 5
        current = write_table(current, "Sabat ke", saturday_dates, data_khotbah, include_header=True, date_label="Tanggal")
        current = write_table(current, "Sabat ke", saturday_dates, data_ss, include_header=False, date_label="Sekolah Sabat")
        current = write_table(current, "Rabu ke", wednesday_dates, data_rabu, include_header=False, date_label="Rabu Malam")
        current = write_table(current, "Jumat ke", friday_dates, data_vesper, include_header=False, date_label="Vesper")

        ws.column_dimensions["A"].width = 26
        for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N"]:
            ws.column_dimensions[col_letter].width = 18

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        resp = make_response(stream.read())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"jadwal_kebaktian_triwulan_{q}_{year}_all.xlsx"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return resp
    # end def

# end class


